import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
from collections import defaultdict

wb = openpyxl.load_workbook(r'c:\Users\alcha\OneDrive\Desktop\Proyectos IA\PRUEBAS RICHI\Consolidado SII.xlsx')
ws_c = wb['Compras 2025']

MESES_ORDER = ['Enero','Febrero','Marzo','Abril','Mayo','Junio','Julio','Agosto','Septiembre','Octubre','Noviembre','Diciembre']
meses = defaultdict(float)
proveedores = defaultdict(float)
tipo_compra = defaultdict(float)
empresas_data = defaultdict(float)
total_neto = total_iva = total_monto = total_exento = count = 0

for row in ws_c.iter_rows(min_row=2, values_only=True):
    if row[0] is None:
        continue
    count += 1
    empresa = row[0]
    mes = row[1]
    neto = row[12] or 0
    iva = row[13] or 0
    total = row[16] or 0
    exento = row[11] or 0
    proveedor = row[6]
    tc = row[4]
    if mes:
        meses[mes] += total
    if proveedor:
        proveedores[proveedor] += total
    if tc:
        tipo_compra[tc] += total
    if empresa:
        empresas_data[empresa] += total
    total_neto += neto
    total_iva += iva
    total_monto += total
    total_exento += exento

top_prov = sorted(proveedores.items(), key=lambda x: x[1], reverse=True)[:10]
meses_sorted = [(m, meses[m]) for m in MESES_ORDER if m in meses]
tipo_list = sorted(tipo_compra.items(), key=lambda x: x[1], reverse=True)
empresas_sorted = sorted(empresas_data.items(), key=lambda x: x[1], reverse=True)

if 'Dashboard' in wb.sheetnames:
    del wb['Dashboard']
ws = wb.create_sheet('Dashboard', 0)

HOT_PINK = 'FF1493'
PINK = 'FF69B4'
LIGHT_PINK = 'FFB6C1'
BLACK = '000000'
CARD = '120010'
DARK = '0D0D0D'

BF = PatternFill('solid', fgColor=BLACK)
CF = PatternFill('solid', fgColor=CARD)
DF = PatternFill('solid', fgColor=DARK)

def sc(cell, fill=None, fc=PINK, bold=False, size=11, ha='center', va='center', wrap=False):
    if fill:
        cell.fill = fill
    else:
        cell.fill = BF
    cell.font = Font(color=fc, bold=bold, size=size, name='Calibri')
    cell.alignment = Alignment(horizontal=ha, vertical=va, wrap_text=wrap)

# Fill all cells black
for r in range(1, 65):
    for c in range(1, 36):
        ws.cell(r, c).fill = BF

# Column widths
ws.column_dimensions['A'].width = 2
for col in ['B','C','D']:
    ws.column_dimensions[col].width = 13
ws.column_dimensions['E'].width = 1
for col in ['F','G','H']:
    ws.column_dimensions[col].width = 13
ws.column_dimensions['I'].width = 1
for col in ['J','K','L']:
    ws.column_dimensions[col].width = 13
ws.column_dimensions['M'].width = 1
for col in ['N','O','P']:
    ws.column_dimensions[col].width = 13
ws.column_dimensions['Q'].width = 1
for col in ['R','S','T']:
    ws.column_dimensions[col].width = 13
ws.column_dimensions['U'].width = 2

# Data cols (for chart references)
ws.column_dimensions['W'].width = 22
ws.column_dimensions['X'].width = 16
ws.column_dimensions['Y'].width = 24
ws.column_dimensions['Z'].width = 16
ws.column_dimensions['AA'].width = 36
ws.column_dimensions['AB'].width = 16

# Row heights
ws.row_dimensions[1].height = 8
ws.row_dimensions[2].height = 40
ws.row_dimensions[3].height = 18
ws.row_dimensions[4].height = 8
ws.row_dimensions[5].height = 16
ws.row_dimensions[6].height = 16
ws.row_dimensions[7].height = 36
ws.row_dimensions[8].height = 14
ws.row_dimensions[9].height = 10
ws.row_dimensions[10].height = 20
for r in range(11, 33):
    ws.row_dimensions[r].height = 16
ws.row_dimensions[33].height = 8
ws.row_dimensions[34].height = 20
for r in range(35, 58):
    ws.row_dimensions[r].height = 16

# ======= TITLE =======
ws.merge_cells('B2:T2')
t = ws['B2']
t.value = '  DASHBOARD DE COMPRAS 2025'
sc(t, BF, HOT_PINK, bold=True, size=24)
t.border = Border(bottom=Side(style='thick', color=HOT_PINK))

ws.merge_cells('B3:T3')
s = ws['B3']
s.value = 'Consolidado SII  |  5 Empresas  |  12 Meses  |  3,593 Documentos'
sc(s, BF, LIGHT_PINK, size=10)
s.border = Border(bottom=Side(style='thin', color='550033'))

# ======= KPI CARDS =======
kpis = [
    ('DOCUMENTOS',      f'{count:,}',                 'B', 'D'),
    ('MONTO NETO',      f'${total_neto/1e6:.2f}M',    'F', 'H'),
    ('IVA RECUPERABLE', f'${total_iva/1e6:.2f}M',     'J', 'L'),
    ('MONTO TOTAL',     f'${total_monto/1e6:.2f}M',   'N', 'P'),
    ('MONTO EXENTO',    f'${total_exento/1e6:.2f}M',  'R', 'T'),
]

for label, value, c1, c2 in kpis:
    ws.merge_cells(f'{c1}5:{c2}6')
    lc = ws[f'{c1}5']
    lc.value = label
    lc.fill = CF
    lc.font = Font(color=LIGHT_PINK, bold=True, size=9, name='Calibri')
    lc.alignment = Alignment(horizontal='center', vertical='center')
    lc.border = Border(
        top=Side(style='medium', color=HOT_PINK),
        left=Side(style='medium', color=HOT_PINK),
        right=Side(style='medium', color=HOT_PINK)
    )

    ws.merge_cells(f'{c1}7:{c2}7')
    vc = ws[f'{c1}7']
    vc.value = value
    vc.fill = CF
    vc.font = Font(color=HOT_PINK, bold=True, size=20, name='Calibri')
    vc.alignment = Alignment(horizontal='center', vertical='center')
    vc.border = Border(
        left=Side(style='medium', color=HOT_PINK),
        right=Side(style='medium', color=HOT_PINK)
    )

    ws.merge_cells(f'{c1}8:{c2}8')
    bc = ws[f'{c1}8']
    bc.fill = CF
    bc.border = Border(
        bottom=Side(style='medium', color=HOT_PINK),
        left=Side(style='medium', color=HOT_PINK),
        right=Side(style='medium', color=HOT_PINK)
    )

# ======= SECTION HEADERS =======
ws.merge_cells('B10:K10')
sh1 = ws['B10']
sh1.value = '  COMPRAS POR MES (Monto Total en Millones CLP)'
sc(sh1, DF, HOT_PINK, bold=True, size=11, ha='left')
sh1.border = Border(
    left=Side(style='thick', color=HOT_PINK),
    bottom=Side(style='thin', color=HOT_PINK)
)

ws.merge_cells('M10:T10')
sh2 = ws['M10']
sh2.value = '  TIPO DE COMPRA'
sc(sh2, DF, HOT_PINK, bold=True, size=11, ha='left')
sh2.border = Border(
    left=Side(style='thick', color=HOT_PINK),
    bottom=Side(style='thin', color=HOT_PINK)
)

ws.merge_cells('B33:T33')
sh3 = ws['B33']
sh3.value = '  TOP 10 PROVEEDORES POR MONTO TOTAL (Millones CLP)'
sc(sh3, DF, HOT_PINK, bold=True, size=11, ha='left')
sh3.border = Border(
    left=Side(style='thick', color=HOT_PINK),
    bottom=Side(style='thin', color=HOT_PINK)
)

# ======= EMPRESA TABLE (rows 11-28, cols U-X) =======
ws.merge_cells('V10:Y10')
eh = ws['V10']
eh.value = '  RESUMEN POR EMPRESA'
sc(eh, DF, HOT_PINK, bold=True, size=11, ha='left')
eh.border = Border(
    left=Side(style='thick', color=HOT_PINK),
    bottom=Side(style='thin', color=HOT_PINK)
)

# Set col widths for empresa table
ws.column_dimensions['V'].width = 28
ws.column_dimensions['W'].width = 14
ws.column_dimensions['X'].width = 10
ws.column_dimensions['Y'].width = 22

# Table header row 11
for col, hdr in zip([22, 23, 24, 25], ['EMPRESA', 'MONTO', '%', 'BARRA']):
    c = ws.cell(11, col)
    c.value = hdr
    c.fill = CF
    c.font = Font(color=HOT_PINK, bold=True, size=9, name='Calibri')
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = Border(
        bottom=Side(style='thin', color=HOT_PINK),
        top=Side(style='thin', color=HOT_PINK)
    )

for row_i, (empresa, monto) in enumerate(empresas_sorted, 12):
    pct = monto / total_monto * 100
    bar_blocks = int(pct / 5)
    bar_str = '█' * bar_blocks + '░' * (20 - bar_blocks)
    row_fill = CF if row_i % 2 == 0 else BF
    row_fc = PINK if row_i % 2 == 0 else LIGHT_PINK

    emp_c = ws.cell(row_i, 22)
    emp_c.value = empresa
    emp_c.fill = row_fill
    emp_c.font = Font(color=row_fc, size=9, name='Calibri')
    emp_c.alignment = Alignment(horizontal='left', vertical='center')

    mnt_c = ws.cell(row_i, 23)
    mnt_c.value = f'${monto/1e6:.2f}M'
    mnt_c.fill = row_fill
    mnt_c.font = Font(color=HOT_PINK, bold=True, size=9, name='Calibri')
    mnt_c.alignment = Alignment(horizontal='center', vertical='center')

    pct_c = ws.cell(row_i, 24)
    pct_c.value = f'{pct:.1f}%'
    pct_c.fill = row_fill
    pct_c.font = Font(color=row_fc, size=9, name='Calibri')
    pct_c.alignment = Alignment(horizontal='center', vertical='center')

    bar_c = ws.cell(row_i, 25)
    bar_c.value = bar_str[:20]
    bar_c.fill = row_fill
    bar_c.font = Font(color=HOT_PINK, size=8, name='Consolas')
    bar_c.alignment = Alignment(horizontal='left', vertical='center')

# ======= DATA FOR CHARTS (col W=23, X=24... wait col mapping) =======
# Using columns 27 (AA), 28 (AB) for meses
# columns 29 (AC), 30 (AD) for tipo
# columns 31 (AE), 32 (AF) for proveedores

ws.column_dimensions['AA'].width = 22
ws.column_dimensions['AB'].width = 14
ws.column_dimensions['AC'].width = 28
ws.column_dimensions['AD'].width = 14
ws.column_dimensions['AE'].width = 36
ws.column_dimensions['AF'].width = 14

# Meses data (col AA=27, AB=28)
ws.cell(1, 27).value = 'Mes'
ws.cell(1, 28).value = 'Monto'
for i, (mes, val) in enumerate(meses_sorted, 2):
    ws.cell(i, 27).value = mes
    ws.cell(i, 28).value = round(val / 1e6, 2)

# Tipo compra data (col AC=29, AD=30)
ws.cell(1, 29).value = 'Tipo'
ws.cell(1, 30).value = 'Monto'
for i, (tipo, val) in enumerate(tipo_list, 2):
    ws.cell(i, 29).value = tipo
    ws.cell(i, 30).value = round(val / 1e6, 2)

# Top prov data (col AE=31, AF=32)
ws.cell(1, 31).value = 'Proveedor'
ws.cell(1, 32).value = 'Monto'
for i, (prov, val) in enumerate(top_prov, 2):
    short = prov[:40] if len(prov) > 40 else prov
    ws.cell(i, 31).value = short
    ws.cell(i, 32).value = round(val / 1e6, 2)

# Style data cells black/pink
for r in range(1, 14):
    for c in range(27, 33):
        cell = ws.cell(r, c)
        cell.fill = BF
        cell.font = Font(color=PINK, size=9)

# ======= CHART 1: Bar - Compras por Mes =======
bar1 = BarChart()
bar1.type = 'col'
bar1.grouping = 'clustered'
bar1.title = None
bar1.style = 26
bar1.y_axis.title = 'M CLP'
bar1.x_axis.title = None
bar1.width = 18
bar1.height = 12

cats1 = Reference(ws, min_col=27, max_col=27, min_row=2, max_row=1 + len(meses_sorted))
data1 = Reference(ws, min_col=28, max_col=28, min_row=1, max_row=1 + len(meses_sorted))
bar1.add_data(data1, titles_from_data=True)
bar1.set_categories(cats1)

try:
    bar1.series[0].graphicalProperties.solidFill = HOT_PINK
    bar1.series[0].graphicalProperties.line.solidFill = 'C71585'
except Exception:
    pass

ws.add_chart(bar1, 'B11')

# ======= CHART 2: Pie - Tipo Compra =======
pie = PieChart()
pie.title = None
pie.style = 26
pie.width = 12
pie.height = 12

cats2 = Reference(ws, min_col=29, max_col=29, min_row=2, max_row=1 + len(tipo_list))
data2 = Reference(ws, min_col=30, max_col=30, min_row=1, max_row=1 + len(tipo_list))
pie.add_data(data2, titles_from_data=True)
pie.set_categories(cats2)

PINK_SHADES = ['FF1493', 'FF69B4', 'C71585', 'FFB6C1', 'FF007F']
for i in range(len(tipo_list)):
    dp = DataPoint(idx=i)
    dp.graphicalProperties.solidFill = PINK_SHADES[i % len(PINK_SHADES)]
    pie.series[0].dPt.append(dp)

ws.add_chart(pie, 'M11')

# ======= CHART 3: Bar horizontal - Top Proveedores =======
bar2 = BarChart()
bar2.type = 'bar'
bar2.grouping = 'clustered'
bar2.title = None
bar2.style = 26
bar2.x_axis.title = 'Millones CLP'
bar2.y_axis.title = None
bar2.width = 32
bar2.height = 14

cats3 = Reference(ws, min_col=31, max_col=31, min_row=2, max_row=1 + len(top_prov))
data3 = Reference(ws, min_col=32, max_col=32, min_row=1, max_row=1 + len(top_prov))
bar2.add_data(data3, titles_from_data=True)
bar2.set_categories(cats3)

try:
    bar2.series[0].graphicalProperties.solidFill = HOT_PINK
    bar2.series[0].graphicalProperties.line.solidFill = 'C71585'
except Exception:
    pass

ws.add_chart(bar2, 'B35')

print('Guardando...')
wb.save(r'c:\Users\alcha\OneDrive\Desktop\Proyectos IA\PRUEBAS RICHI\Consolidado SII.xlsx')
print('Dashboard creado exitosamente!')
